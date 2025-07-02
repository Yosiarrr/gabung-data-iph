import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import zipfile
from datetime import datetime
import xlwt

st.title("📊 Aplikasi Gabung Data Excel Harga IPH")

# Pilih Tahun & Bulan
tahun = st.selectbox("Pilih Tahun", options=[2023, 2024, 2025], index=2)
bulan_nama = st.selectbox(
    "Pilih Bulan",
    options=[
        "Januari", "Februari", "Maret", "April", "Mei", "Juni",
        "Juli", "Agustus", "September", "Oktober", "November", "Desember"
    ],
    index=0
)

map_bulan = {
    "Januari": "01", "Februari": "02", "Maret": "03", "April": "04",
    "Mei": "05", "Juni": "06", "Juli": "07", "Agustus": "08",
    "September": "09", "Oktober": "10", "November": "11", "Desember": "12"
}
bulan = map_bulan[bulan_nama]

uploaded_files = st.file_uploader("Upload beberapa file Excel (.xlsx)", type=["xlsx"], accept_multiple_files=True)

# Fungsi ekstrak minggu dari nama file
def extract_minggu(filename):
    for i in range(1, 6):
        if f"M{i}" in filename.upper():
            return i
    return None

if st.button("Proses & Unduh .zip") and uploaded_files:
    semua_data_prov = []
    semua_data_kab = []

    if tahun == 2025:
        indeks_kolom = [0, 1, 2, 3, 4, 5]
    elif tahun == 2024:
        indeks_kolom = [0, 1, 2, 3, 4, 5]
    else:
        indeks_kolom = [0, 1, 2, 3, 4, 5]

    for uploaded_file in uploaded_files:
        try:
            wb = load_workbook(uploaded_file, data_only=True)
            nama_file = uploaded_file.name
            minggu = extract_minggu(nama_file)

            # PROVINSI
            if "Provinsi" in wb.sheetnames:
                sheet_prov = wb["Provinsi"]
                for row in sheet_prov.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        selected = [row[i] if i < len(row) else None for i in indeks_kolom]
                        semua_data_prov.append((minggu, selected))

            # KABUPATEN
            if "360 KabKota" in wb.sheetnames:
                sheet_kab = wb["360 KabKota"]
                for row in sheet_kab.iter_rows(min_row=2, values_only=True):
                    if row[0] and str(row[0]).startswith("18"):
                        selected = [row[i] if i < len(row) else None for i in indeks_kolom]
                        semua_data_kab.append((minggu, selected))

        except Exception as e:
            st.error(f"❌ Gagal memproses file {uploaded_file.name}: {e}")

    if semua_data_prov or semua_data_kab:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            today = datetime.today().strftime("%Y-%m-%d")

            # GABUNG PROVINSI
            if semua_data_prov:
                book_prov = xlwt.Workbook()
                sheet_prov = book_prov.add_sheet("Gabungan_Provinsi")
                headers = [
                    "id", "tahun", "bulan", "minggu", "kode_prov", "prov", "nilai_iph",
                    "komoditas", "fluktuasi_harga_tertinggi", "nilai_fluktuasi_tertinggi",
                    "disparitas_harga_antar_wilayah", "date_created"
                ]
                for col, val in enumerate(headers):
                    sheet_prov.write(0, col, val)
                for idx, (minggu, row) in enumerate(semua_data_prov, start=1):
                    komoditas = str(row[3]).replace(",", ";")
                    baris = [
                        idx, str(tahun), bulan, minggu,
                        row[0], row[1], row[2], komoditas,
                        row[4], row[5], "", today
                    ]
                    for col, val in enumerate(baris):
                        sheet_prov.write(idx, col, val)
                output_prov = io.BytesIO()
                book_prov.save(output_prov)
                output_prov.seek(0)
                zip_file.writestr(f"gabungan_{bulan}_{tahun}_provinsi.xls", output_prov.read())

            # GABUNG KABUPATEN
            if semua_data_kab:
                book_kab = xlwt.Workbook()
                sheet_kab = book_kab.add_sheet("Gabungan_Kabupaten")
                headers = [
                    "id", "tahun", "bulan", "minggu", "kode_kab", "kab", "nilai_iph",
                    "komoditas", "fluktuasi_harga_tertinggi", "nilai_fluktuasi_tertinggi",
                    "disparitas_harga_antar_wilayah", "date_created"
                ]
                for col, val in enumerate(headers):
                    sheet_kab.write(0, col, val)
                for idx, (minggu, row) in enumerate(semua_data_kab, start=1):
                    komoditas = str(row[3]).replace(",", ";")
                    baris = [
                        idx, str(tahun), bulan, minggu,
                        row[0], row[1], row[2], komoditas,
                        row[4], row[5], "", today
                    ]
                    for col, val in enumerate(baris):
                        sheet_kab.write(idx, col, val)
                output_kab = io.BytesIO()
                book_kab.save(output_kab)
                output_kab.seek(0)
                zip_file.writestr(f"gabungan_{bulan}_{tahun}_kabupaten.xls", output_kab.read())

        zip_buffer.seek(0)
        st.success("✅ Data berhasil diproses!")
        st.download_button(
            label="📥 Unduh Gabungan File (.zip)",
            data=zip_buffer,
            file_name=f"gabungan_IPH_{bulan}_{tahun}.zip",
            mime="application/zip"
        )
    else:
        st.warning("❗ Tidak ada data yang diproses.")
